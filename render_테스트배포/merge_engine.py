# -*- coding: utf-8 -*-
"""멀티문서 병합 엔진 (나스 배포 / 자동병합 / 최종본 다운로드)."""
import os, datetime, shutil, glob
from copy import copy
import openpyxl

BASE = os.path.dirname(os.path.abspath(__file__))
# 데이터 루트: 나스 공유폴더로 매핑 (Docker 볼륨). 미지정 시 로컬 data/
ROOT = os.environ.get("DATA_DIR", os.path.join(BASE, "data"))
TEMPLATES = os.path.join(ROOT, "00_양식")
INBOX_ROOT = os.path.join(ROOT, "inbox")
MASTER_ROOT = os.path.join(ROOT, "90_병합본")
MONTH = os.environ.get("MONTH", "26년 5월")

ENTITIES = ["KOGK", "TOGK", "QHANO", "HOGK", "Shunran"]
DOCUMENTS = {
    "경영위원회": {"label": "경영위원회 자료", "mode": "by_entity",
                "template": "경영위원회.xlsx", "uploaders": ENTITIES},
    "KPI": {"label": "KPI Reporting (Onyx)", "mode": "by_entity",
            "template": "KPI.xlsx", "uploaders": ENTITIES},
    "연결정산표": {"label": "연결정산표", "mode": "single",
                "template": None, "uploaders": ["연결담당"]},
}


def ensure():
    for k in DOCUMENTS:
        os.makedirs(os.path.join(INBOX_ROOT, k), exist_ok=True)
    os.makedirs(TEMPLATES, exist_ok=True)
    os.makedirs(MASTER_ROOT, exist_ok=True)


def inbox_dir(doc_key):
    d = os.path.join(INBOX_ROOT, doc_key)
    os.makedirs(d, exist_ok=True)
    return d


def owned_sheets(wb, entity):
    out = []
    for sn in wb.sheetnames:
        hit = [e for e in ENTITIES if e.lower() in sn.lower()]
        if len(hit) == 1 and hit[0] == entity:
            out.append(sn)
    return out


def copy_sheet(ssrc, sdst):
    for mr in list(sdst.merged_cells.ranges):
        sdst.unmerge_cells(str(mr))
    for row in ssrc.iter_rows():
        for c in row:
            if c.value is None and not c.has_style:
                continue
            d = sdst[c.coordinate]
            d.value = c.value
            if c.has_style:
                d._style = copy(c._style)
    for mr in ssrc.merged_cells.ranges:
        sdst.merge_cells(str(mr))
    for k, dim in ssrc.column_dimensions.items():
        if dim.width is not None:
            sdst.column_dimensions[k].width = dim.width
    for k, dim in ssrc.row_dimensions.items():
        if dim.height is not None:
            sdst.row_dimensions[k].height = dim.height


def detect_submissions(doc_key):
    doc = DOCUMENTS[doc_key]
    d = inbox_dir(doc_key)
    found = {}
    for fn in os.listdir(d):
        if not fn.lower().endswith(".xlsx") or fn.startswith("~$"):
            continue
        if doc["mode"] == "single":
            found[doc["uploaders"][0]] = os.path.join(d, fn)
        else:
            for e in doc["uploaders"]:
                if e.lower() in fn.lower():
                    found[e] = os.path.join(d, fn)
                    break
    return found


def latest_master(doc_key):
    files = glob.glob(os.path.join(MASTER_ROOT, doc_key + "_병합본_*.xlsx"))
    if not files:
        return None
    return os.path.basename(max(files, key=os.path.getmtime))


def status_doc(doc_key):
    doc = DOCUMENTS[doc_key]
    subs = detect_submissions(doc_key)
    rows = []
    for u in doc["uploaders"]:
        up = u in subs
        rows.append({"uploader": u, "uploaded": up})
    done = 0
    for r in rows:
        if r["uploaded"]:
            done += 1
    total = len(doc["uploaders"])
    lm = latest_master(doc_key)
    return {"key": doc_key, "label": doc["label"], "mode": doc["mode"],
            "rows": rows, "done": done, "total": total,
            "all_done": done == total, "master": lm}


def status_all():
    docs = []
    for k in DOCUMENTS:
        docs.append(status_doc(k))
    done_docs = 0
    for d in docs:
        if d["all_done"]:
            done_docs += 1
    return {"docs": docs, "done_docs": done_docs, "total_docs": len(docs),
            "all_done": done_docs == len(docs), "month": MONTH}


def run_merge_doc(doc_key):
    doc = DOCUMENTS[doc_key]
    subs = detect_submissions(doc_key)
    os.makedirs(MASTER_ROOT, exist_ok=True)
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    out = os.path.join(MASTER_ROOT, doc_key + "_병합본_" + MONTH.replace(" ", "") + "_" + ts + ".xlsx")
    cells = 0
    if doc["mode"] == "single":
        shutil.copy(list(subs.values())[0], out)
    else:
        tmpl = os.path.join(TEMPLATES, doc["template"])
        master = openpyxl.load_workbook(tmpl)
        for entity in doc["uploaders"]:
            if entity not in subs:
                continue
            src = openpyxl.load_workbook(subs[entity], data_only=False)
            for s in owned_sheets(src, entity):
                if s in master.sheetnames:
                    copy_sheet(src[s], master[s])
            src.close()
        master.save(out)
        master.close()
    return os.path.basename(out)
